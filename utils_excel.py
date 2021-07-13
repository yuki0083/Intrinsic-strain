import openpyxl
import calculation
import os


def write_list_2d(sheet, list_2d, start_row, start_col):
    for y, row in enumerate(list_2d): #y:行番号
        for x, cell in enumerate(row): #x:列番号
            sheet.cell(row = start_row + y,column=start_col + x, value=list_2d[y][x])

def write_excel_sheet(intrinsic_path_list, input_directory_path, col_num, col_name, xyz_path ='./xyz-coordinate.csv',  h=0.7, E=175000.0):
    wb = openpyxl.Workbook()  # ファイル新規作成
    # ws = wb.active
    ws = wb.worksheets[0]  # シートの指定

    #df_xyz(ソリッド要素番号,中心(X),中心(Y),中心(Z),Δx,Δy,Δz)
    df_xyz = calculation.make_df_xyz(xyz_path)

    for i, intrinsic_path in enumerate(intrinsic_path_list):
        #df_intrinsic(ソリッド要素番号,塑性ひずみ(X),塑性ひずみ(Y),塑性ひずみ(Z),塑性ひずみ(XY), 塑性ひずみ(YZ), 塑性ひずみ(ZX))
        df_intrinsic = calculation.make_df_intrisic(intrinsic_path, df_xyz)
        
        file_name, ext = os.path.splitext(os.path.basename(intrinsic_path))

        start_col = 1 + col_num * i + i

        ws.cell(row=1, column=start_col, value=file_name)
        write_list_2d(ws, col_name, start_row=2, start_col=start_col)
        write_list_2d(ws, results, start_row=3, start_col=start_col)

    directory_name = os.path.basename(input_directory_path)
    result_file_name = directory_name + '_' + 'result.xlsx'
    save_path = os.path.join(input_directory_path, result_file_name)
    wb.save(save_path)